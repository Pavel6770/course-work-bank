import openpyxl
from typing import List, Dict, Any


def load_transactions_from_excel(file_path: str) -> List[Dict[str, Any]]:
    """Загружает транзакции из Excel-файла."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Определяем индексы колонок
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = idx

    transactions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_full = str(row[headers["Дата операции"] - 1]) if "Дата операции" in headers else ""
        date_clean = date_full.split()[0] if date_full else ""
        amount = row[headers["Сумма операции"] - 1] if "Сумма операции" in headers else 0
        category = row[headers["Категория"] - 1] if "Категория" in headers else "Без категории"
        description = row[headers["Описание"] - 1] if "Описание" in headers else ""
        cashback = row[headers["Кэшбэк"] - 1] if "Кэшбэк" in headers else 0
        card_number = row[headers["Номер карты"] - 1] if "Номер карты" in headers else ""

        transactions.append(
            {
                "date": date_clean,
                "amount": float(amount) if amount else 0,
                "category": str(category) if category else "Без категории",
                "description": str(description) if description else "",
                "cashback": float(cashback) if cashback else 0,
                "card_number": str(card_number) if card_number else "",
            }
        )

    return transactions


def get_transactions() -> List[Dict[str, Any]]:
    """Получает транзакции (из файла, потом можно заменить на API/БД)."""
    return load_transactions_from_excel("data/operations.xlsx")


def get_test_transactions() -> List[Dict[str, Any]]:
    """Возвращает тестовые данные (для юнит-тестов)."""
    return [
        {
            "date": "01.01.2024",
            "amount": -1000,
            "category": "Продукты",
            "description": "Магнит",
            "cashback": 50,
            "card_number": "*1234",
        },
        {
            "date": "02.01.2024",
            "amount": -500,
            "category": "Кафе",
            "description": "Кофе",
            "cashback": 25,
            "card_number": "*1234",
        },
        {
            "date": "03.01.2024",
            "amount": 50000,
            "category": "Зарплата",
            "description": "Аванс",
            "cashback": 0,
            "card_number": "*1234",
        },
    ]
