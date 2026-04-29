import openpyxl
from datetime import datetime
from typing import List, Dict, Any


def load_transactions_from_excel(file_path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = idx
    
    transactions = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_full = str(row[headers["Дата операции"] - 1]) if "Дата операции" in headers else ""
        date_clean = date_full.split()[0] if date_full else ""
        
        amount = row[headers["Сумма операции"] - 1] if "Сумма операции" in headers else 0
        category = row[headers["Категория"] - 1] if "Категория" in headers else "Без категории"
        description = row[headers["Описание"] - 1] if "Описание" in headers else ""
        cashback = row[headers["Кэшбэк"] - 1] if "Кэшбэк" in headers else 0
        card_number = row[headers["Номер карты"] - 1] if "Номер карты" in headers else ""
        
        try:
            date_obj = datetime.strptime(date_clean, "%d.%m.%Y")
            formatted_date = date_obj.strftime("%d.%m.%Y")
        except:
            formatted_date = date_clean
        
        transactions.append({
            "date": formatted_date,
            "amount": float(amount) if amount else 0,
            "category": str(category) if category else "Без категории",
            "description": str(description) if description else "",
            "cashback": float(cashback) if cashback else 0,
            "card_number": str(card_number) if card_number else ""
        })
    
    return transactions


def sort_transactions_by_amount(transactions: List[Dict[str, Any]], reverse: bool = True) -> List[Dict[str, Any]]:
    return sorted(transactions, key=lambda x: x["amount"], reverse=reverse)


def get_total_by_category(transactions: List[Dict[str, Any]]) -> Dict[str, float]:
    result = {}
    for t in transactions:
        category = t["category"]
        amount = abs(t["amount"])
        result[category] = result.get(category, 0) + amount
    return result


if __name__ == "__main__":
    trans = load_transactions_from_excel("data/operations.xlsx")
    print(f"Loaded {len(trans)} transactions")
    for t in trans[:3]:
        print(f"  {t['date']} | {t['amount']} | {t['category']}")
